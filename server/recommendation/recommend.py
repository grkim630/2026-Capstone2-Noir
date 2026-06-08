"""
얼굴 분석 JSON + noir_color_db.xlsx → recommendation_result.json 생성.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .color_utils import (
    normalize_hex,
    resolve_product_rgb,
    rgb_to_hsv,
    rgb_to_lab,
    to_rgb01,
)
from .score_utils import (
    TYPE_TEMPERATURE,
    VALID_PERSONAL_COLOR_TYPES,
    calculate_final_score,
    classify_brightness_from_hsv,
    classify_chroma_from_hsv,
    generate_reason,
    sort_recommendation_items,
)

SCORING_VERSION = "mvp_v1.4"
SOURCE_DB_NAME = "noir_color_db.xlsx"

EYE_PALETTE_SHADE_TYPES = {
    "eye_palette_glitter": "glitter",
    "eye_palette_base_matt": "base_matt",
    "eye_palette_semi_glitter_matt": "semi_glitter_matt",
    "eye_palette_blending_matt": "blending_matt",
}
REQUIRED_EYE_SHADES = list(EYE_PALETTE_SHADE_TYPES.values())

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_DB_PATHS = [
    PROJECT_ROOT / "data" / SOURCE_DB_NAME,
    PROJECT_ROOT / "server" / "data" / SOURCE_DB_NAME,
]

COLUMN_ALIASES = {
    "product type": "productType",
    "id": "id",
    "name": "name",
    "hex": "hex",
    "rgb": "rgb",
    "personal color (8 types)": "personalColor",
}


def normalize_product_type(value: str) -> tuple[str | None, str | None]:
    """반환: (카테고리, 아이팔레트 shade 키). shade는 eye_palette_* 행에만 존재."""
    normalized = str(value or "").strip().lower().replace(" ", "_")
    if normalized in ("brush",):
        return None, None
    if normalized in EYE_PALETTE_SHADE_TYPES:
        return "eye_palette", EYE_PALETTE_SHADE_TYPES[normalized]
    if normalized in ("lip", "blush", "eye_palette"):
        return normalized, None
    if normalized == "eye":
        return "eye_palette", None
    return None, None


def format_palette_display_name(name: str) -> str:
    return " ".join(part.capitalize() for part in str(name).replace("_", " ").split())


def average_rgb(rgb_values: list[list[int]]) -> list[int]:
    if not rgb_values:
        return [0, 0, 0]
    length = len(rgb_values[0])
    return [
        round(sum(row[index] for row in rgb_values) / len(rgb_values))
        for index in range(length)
    ]


def resolve_db_path(db_path: Path | str | None = None) -> Path:
    if db_path:
        path = Path(db_path)
        if not path.exists():
            raise FileNotFoundError(f"Cosmetic DB not found: {path}")
        return path

    for candidate in DEFAULT_DB_PATHS:
        if candidate.exists():
            return candidate

    raise FileNotFoundError(
        f"Cosmetic DB not found. Place {SOURCE_DB_NAME} at: {DEFAULT_DB_PATHS[0]}",
    )


def load_cosmetic_db(file_path: Path | str) -> dict[str, Any]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active

    rows = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration as error:
        raise ValueError("Cosmetic DB sheet is empty.") from error

    column_map: dict[int, str] = {}
    for index, header in enumerate(header_row):
        if header is None:
            continue
        key = COLUMN_ALIASES.get(str(header).strip().lower())
        if key:
            column_map[index] = key

    required = {"productType", "id", "name", "hex", "rgb", "personalColor"}
    if set(column_map.values()) < required:
        raise ValueError(
            "Cosmetic DB header must include: Product Type, ID, Name, HEX, RGB, Personal Color (8 Types)",
        )

    products: list[dict[str, Any]] = []
    eye_palette_groups: dict[int, dict[str, Any]] = {}

    for row in rows:
        if not row or all(cell is None for cell in row):
            continue

        record: dict[str, Any] = {}
        for index, key in column_map.items():
            record[key] = row[index] if index < len(row) else None

        product_type, shade_key = normalize_product_type(record.get("productType"))
        if not product_type:
            continue

        personal_color = str(record.get("personalColor") or "").strip()
        if personal_color not in VALID_PERSONAL_COLOR_TYPES:
            continue

        try:
            hex_value = normalize_hex(str(record.get("hex") or ""))
            rgb = resolve_product_rgb(hex_value, str(record.get("rgb") or ""))
        except (ValueError, TypeError):
            continue

        try:
            product_id = int(record.get("id"))
        except (TypeError, ValueError):
            continue

        name = str(record.get("name") or "").strip()
        if not name:
            continue

        if product_type == "eye_palette" and shade_key:
            group = eye_palette_groups.setdefault(
                product_id,
                {
                    "productType": "eye_palette",
                    "id": product_id,
                    "name": name,
                    "shades": {},
                },
            )
            if group["name"] != name:
                group["name"] = name
            group["shades"][shade_key] = {
                "shade": shade_key,
                "hex": hex_value,
                "rgb": rgb,
                "personalColor": personal_color,
            }
            continue

        product_record: dict[str, Any] = {
                "productType": product_type,
                "id": product_id,
                "name": name,
                "hex": hex_value,
                "rgb": rgb,
                "personalColor": personal_color,
            }
        if product_type in ("lip", "blush"):
            product_record["displayName"] = format_display_name(name)
            product_record["displayID"] = format_display_id(product_id, 3)
        products.append(product_record)

    workbook.close()

    eye_palettes: list[dict[str, Any]] = []
    for product_id in sorted(eye_palette_groups):
        group = eye_palette_groups[product_id]
        shades = group["shades"]
        if not all(shade in shades for shade in REQUIRED_EYE_SHADES):
            continue

        shade_rgbs = [shades[shade]["rgb"] for shade in REQUIRED_EYE_SHADES]
        representative_rgb = average_rgb(shade_rgbs)
        base_shade = shades["base_matt"]
        personal_color = base_shade["personalColor"]

        eye_palettes.append(
            {
                "productType": "eye_palette",
                "id": product_id,
                "name": group["name"],
                "displayName": format_palette_display_name(group["name"]),
                "displayID": format_display_id(product_id, 2),
                "hex": base_shade["hex"],
                "rgb": representative_rgb,
                "personalColor": personal_color,
                "shades": shades,
            },
        )

    return {
        "products": products,
        "eyePalettes": eye_palettes,
    }


def load_face_analysis(file_path: Path | str) -> dict[str, Any]:
    path = Path(file_path)
    with path.open("r", encoding="utf-8") as file:
        data = json.load(file)

    if "front" in data:
        return data

    if isinstance(data, dict) and "skinAnalysis" in data:
        return {"front": data}

    raise ValueError("Face analysis JSON must contain `front` or be a front pose object.")


def build_user_analysis(face_analysis: dict[str, Any]) -> dict[str, Any]:
    front = face_analysis["front"]
    skin_analysis = front.get("skinAnalysis") or {}
    tone = skin_analysis.get("tone") or {}
    lab = skin_analysis.get("LAB") or {}
    hsv = skin_analysis.get("HSV") or {}
    personal_color = front.get("personalColor") or {}

    def color_block(key: str) -> dict[str, Any]:
        color = front.get(key) or {}
        return {
            "rgb": color.get("rgb"),
            "hex": color.get("hex"),
        }

    return {
        "season": personal_color.get("season"),
        "temperature": tone.get("temperature"),
        "brightness": tone.get("brightness"),
        "chroma": tone.get("chroma"),
        "personalColorBrightness": personal_color.get("brightness"),
        "personalColorChroma": personal_color.get("chroma"),
        "skinRGB": front.get("skinColor", {}).get("rgb"),
        "skinHEX": front.get("skinColor", {}).get("hex"),
        "skinLAB": {
            "L": lab.get("L"),
            "a": lab.get("a"),
            "b": lab.get("b"),
        },
        "skinHSV": {
            "H": hsv.get("H"),
            "S": hsv.get("S"),
            "V": hsv.get("V"),
        },
        "skinBaseColor": skin_analysis.get("skinBaseColor") or front.get("skinBaseColor"),
        "makeup": skin_analysis.get("makeup") or front.get("makeup"),
        "quality": front.get("quality"),
        "melanin": skin_analysis.get("melanin"),
        "redness": skin_analysis.get("redness"),
        "facialColors": {
            "skin": color_block("skinColor"),
            "lip": color_block("lipColor"),
            "cheek": color_block("cheekColor"),
            "forehead": color_block("foreheadColor"),
            "eyeArea": color_block("eyeAreaColor"),
            "iris": color_block("irisColor"),
        },
    }


def build_product_color_analysis(product: dict[str, Any]) -> dict[str, Any]:
    hsv = rgb_to_hsv(product["rgb"])
    lab = rgb_to_lab(product["rgb"])
    return {
        "hsv": hsv,
        "lab": lab,
        "brightness": classify_brightness_from_hsv(hsv),
        "chroma": classify_chroma_from_hsv(hsv),
        "temperature": TYPE_TEMPERATURE.get(product["personalColor"]),
    }


def build_touch_designer_mapping(
    product: dict[str, Any],
    *,
    shade_key: str | None = None,
) -> dict[str, Any]:
    product_type = product["productType"]
    product_id = product["id"]
    rgb = product["rgb"]
    hex_value = product["hex"]
    product_id_suffix = f"{product_id}_{shade_key}" if shade_key else str(product_id)

    return {
        "mapType": product_type,
        "productId": f"{product_type}_{product_id_suffix}",
        "colorHEX": hex_value,
        "colorRGB": rgb,
        "colorRGB01": to_rgb01(rgb),
        "intensity": 1.0,
        "blendMode": "multiply_or_overlay",
    }


def build_shade_recommendations(product: dict[str, Any]) -> dict[str, Any]:
    shades_out: dict[str, Any] = {}
    palette_shades = product.get("shades") or {}
    for shade_key in REQUIRED_EYE_SHADES:
        shade = palette_shades.get(shade_key)
        if not shade:
            continue
        shade_product = {
            **product,
            "hex": shade["hex"],
            "rgb": shade["rgb"],
            "personalColor": shade["personalColor"],
        }
        shades_out[shade_key] = {
            "shade": shade_key,
            "hex": shade["hex"],
            "rgb": shade["rgb"],
            "personalColor": shade["personalColor"],
            "productColorAnalysis": build_product_color_analysis(shade_product),
            "touchDesigner": build_touch_designer_mapping(shade_product, shade_key=shade_key),
        }
    return shades_out


def format_display_name(name: str) -> str:
    return str(name or "").replace("_", " ").upper()


def format_display_id(product_id: int | str, digits: int) -> str:
    return str(product_id).zfill(digits)


def build_product_recommendation(user_analysis: dict[str, Any], product: dict[str, Any]) -> dict[str, Any]:
    scores = calculate_final_score(user_analysis, product)
    product_type = product["productType"]
    item = {
        "productType": product_type,
        "id": product["id"],
        "name": product["name"],
        "hex": product["hex"],
        "rgb": product["rgb"],
        "personalColor": product["personalColor"],
        "productColorAnalysis": build_product_color_analysis(product),
        "scores": scores,
        "touchDesigner": build_touch_designer_mapping(product),
    }
    if product_type in ("lip", "blush"):
        item["displayName"] = product.get("displayName") or format_display_name(product["name"])
        item["displayID"] = product.get("displayID") or format_display_id(product["id"], 3)
    elif product_type == "eye_palette":
        item["displayID"] = product.get("displayID") or format_display_id(product["id"], 2)
        item["displayName"] = product.get("displayName") or format_display_name(product["name"])
        palette_shades = build_shade_recommendations(product)
        if len(palette_shades) != len(REQUIRED_EYE_SHADES):
            raise ValueError(
                f"Eye palette id={product['id']} is missing shades. "
                f"Expected {REQUIRED_EYE_SHADES}, got {list(palette_shades)}",
            )
        item["shades"] = palette_shades
    item["reason"] = generate_reason(user_analysis, item)
    return item


def recommend_by_category(
    user_analysis: dict[str, Any],
    products: list[dict[str, Any]],
    category: str,
    limit: int,
) -> list[dict[str, Any]]:
    category_products = [product for product in products if product["productType"] == category]
    scored = [build_product_recommendation(user_analysis, product) for product in category_products]
    ranked = sort_recommendation_items(scored)[:limit]

    if limit > 1:
        for index, item in enumerate(ranked, start=1):
            item["categoryRank"] = index

    return ranked


def generate_recommendation_json(
    user_analysis: dict[str, Any],
    lip_recommendations: list[dict[str, Any]],
    blush_recommendations: list[dict[str, Any]],
    eye_palette_recommendations: list[dict[str, Any]],
    *,
    source_face_analysis: str,
    session_id: str | None = None,
) -> dict[str, Any]:
    payload = {
        "userAnalysis": user_analysis,
        "recommendationSummary": {
            "totalCount": 7,
            "categoryCounts": {
                "lip": 3,
                "blush": 3,
                "eye_palette": 1,
            },
            "rankingRule": (
                "Lip and blush include categoryRank (1–3 within each category). "
                "Eye palette has a single pick with no rank."
            ),
        },
        "recommendations": {
            "lip": lip_recommendations,
            "blush": blush_recommendations,
            "eye_palette": eye_palette_recommendations,
        },
        "metadata": {
            "sourceDB": SOURCE_DB_NAME,
            "sourceFaceAnalysis": source_face_analysis,
            "outputFile": "recommendation_result.json",
            "scoringVersion": SCORING_VERSION,
            "eyePaletteShadeKeys": REQUIRED_EYE_SHADES,
            "description": (
                "Personal color and facial tone based cosmetic recommendation result "
                "for TouchDesigner mapping. eye_palette entries include all four shades."
            ),
        },
    }
    if session_id:
        return {"sessionId": session_id, **payload}
    return payload


def save_recommendation_json(output_path: Path | str, result: dict[str, Any]) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as file:
        json.dump(result, file, indent=2, ensure_ascii=False)
    return path


def run_recommendation(
    face_analysis: dict[str, Any],
    *,
    db_path: Path | str | None = None,
    source_face_analysis_name: str = "final_result.json",
    session_id: str | None = None,
) -> dict[str, Any]:
    db_file = resolve_db_path(db_path)
    catalog = load_cosmetic_db(db_file)
    products = catalog["products"]
    eye_palettes = catalog["eyePalettes"]
    user_analysis = build_user_analysis(face_analysis)

    if not user_analysis.get("season"):
        raise ValueError("Missing front.personalColor.season in face analysis JSON.")

    lip_recommendations = recommend_by_category(user_analysis, products, "lip", 3)
    blush_recommendations = recommend_by_category(user_analysis, products, "blush", 3)
    eye_palette_recommendations = recommend_by_category(user_analysis, eye_palettes, "eye_palette", 1)

    selected = lip_recommendations + blush_recommendations + eye_palette_recommendations
    if len(selected) != 7:
        counts = {
            "lip": len([p for p in products if p["productType"] == "lip"]),
            "blush": len([p for p in products if p["productType"] == "blush"]),
            "eye_palette": len(eye_palettes),
        }
        raise ValueError(
            "Recommendation count is not 7. "
            f"Selected={len(selected)}. DB category counts={counts}. "
            "Check product DB category counts.",
        )

    resolved_session_id = session_id or face_analysis.get("sessionId")

    return generate_recommendation_json(
        user_analysis,
        lip_recommendations,
        blush_recommendations,
        eye_palette_recommendations,
        source_face_analysis=source_face_analysis_name,
        session_id=resolved_session_id,
    )


def run_recommendation_for_session(
    session_dir: Path | str,
    face_analysis: dict[str, Any] | None = None,
    *,
    db_path: Path | str | None = None,
) -> Path:
    session_path = Path(session_dir)
    face_path = session_path / "final_result.json"

    if face_analysis is None:
        face_analysis = load_face_analysis(face_path)

    result = run_recommendation(
        face_analysis,
        db_path=db_path,
        source_face_analysis_name=face_path.name,
        session_id=session_path.name,
    )
    output_path = session_path / "recommendation_result.json"
    save_recommendation_json(output_path, result)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate recommendation_result.json")
    parser.add_argument(
        "--face",
        required=True,
        help="Path to final_result.json or face_analysis.json",
    )
    parser.add_argument(
        "--db",
        default=None,
        help=f"Path to {SOURCE_DB_NAME} (default: data/{SOURCE_DB_NAME})",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output path (default: same folder as face analysis file)",
    )
    args = parser.parse_args()

    face_path = Path(args.face)
    face_analysis = load_face_analysis(face_path)
    session_id = face_analysis.get("sessionId")
    if not session_id and face_path.parent.name.startswith("session_"):
        session_id = face_path.parent.name
    result = run_recommendation(
        face_analysis,
        db_path=args.db,
        source_face_analysis_name=face_path.name,
        session_id=session_id,
    )

    output_path = Path(args.output) if args.output else face_path.parent / "recommendation_result.json"
    save_recommendation_json(output_path, result)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    main()
